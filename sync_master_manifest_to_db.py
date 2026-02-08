#!/usr/bin/env python3
"""
ETL script: Parse TechLiquidators XLSX manifest files into PostgreSQL tl_manifest_items table.

Reads all order_manifest_*.xlsx files from the manifests directory, cross-references
orders.json for order metadata (date, COGS allocation), and upserts individual product
rows into the tl_manifest_items table.

Usage:
  python3 sync_master_manifest_to_db.py [--manifests-dir PATH] [--orders-json PATH]
"""

import os
import sys
import json
import glob
import argparse
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("ERROR: psycopg2 not installed. Run: pip3 install psycopg2-binary")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Database setup
# ---------------------------------------------------------------------------

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS tl_manifest_items (
    id SERIAL PRIMARY KEY,
    order_id VARCHAR(20) NOT NULL,
    listing_id VARCHAR(20),
    listing_title TEXT,
    category VARCHAR(255),
    product_name TEXT,
    upc VARCHAR(20),
    asin VARCHAR(20),
    quantity INTEGER DEFAULT 1,
    unit_retail DOUBLE PRECISION,
    total_retail DOUBLE PRECISION,
    order_date VARCHAR(50),
    line_item_brands VARCHAR(500),
    allocated_cogs_per_unit DOUBLE PRECISION DEFAULT 0,
    created_at TIMESTAMP DEFAULT NOW(),
    UNIQUE(order_id, listing_id, product_name, upc)
);

CREATE INDEX IF NOT EXISTS idx_tl_manifest_items_order_id ON tl_manifest_items(order_id);
CREATE INDEX IF NOT EXISTS idx_tl_manifest_items_upc ON tl_manifest_items(upc);
CREATE INDEX IF NOT EXISTS idx_tl_manifest_items_category ON tl_manifest_items(category);
"""


def get_db_connection():
    """Connect to PostgreSQL using DATABASE_URL_COGS env var."""
    db_url = os.environ.get("DATABASE_URL_COGS") or os.environ.get("DATABASE_URL")
    if not db_url:
        print("ERROR: Set DATABASE_URL_COGS or DATABASE_URL env var")
        sys.exit(1)
    return psycopg2.connect(db_url)


# ---------------------------------------------------------------------------
# Orders.json parser – get order metadata
# ---------------------------------------------------------------------------

def load_orders(orders_json_path: str) -> dict:
    """Load orders.json and return dict keyed by order_id."""
    if not os.path.exists(orders_json_path):
        print(f"WARNING: orders.json not found at {orders_json_path}")
        return {}

    with open(orders_json_path, "r") as f:
        data = json.load(f)

    orders_map = {}
    for order in data.get("orders", []):
        oid = order.get("order_id", "")
        if not oid:
            continue

        # Calculate total order cost (sum of item prices) for COGS allocation
        items = order.get("items", [])
        total_cost = sum(item.get("price", 0) for item in items)
        total_msrp = sum(item.get("msrp", 0) for item in items)
        total_item_count = sum(item.get("item_count", 0) for item in items)

        # Build pallet_id -> line_item mapping for brand extraction
        pallet_brands = {}
        for item in items:
            title = item.get("title", "")
            # Extract brands from title: "Vacuums & Floorcare - iRobot, BISSELL, Shark - Orig. Retail $28,292"
            parts = title.split(" - ")
            brands = parts[1] if len(parts) >= 3 else ""
            for pid in item.get("pallet_ids", []):
                pallet_brands[pid] = brands

        orders_map[oid] = {
            "date": order.get("date", ""),
            "total_cost": total_cost,
            "total_msrp": total_msrp,
            "total_item_count": total_item_count,
            "pallet_brands": pallet_brands,
            "items": items,
        }

    return orders_map


# ---------------------------------------------------------------------------
# XLSX manifest parser
# ---------------------------------------------------------------------------

EXPECTED_HEADERS = [
    "listing id", "listing title", "category", "product name",
    "upc", "asin", "quantity", "orig. retail", "total orig. retail", "stock image"
]


def parse_manifest_xlsx(filepath: str, order_id: str, order_meta: dict) -> list:
    """Parse a single XLSX manifest file and return list of row dicts."""
    rows = []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR reading {filepath}: {e}")
        return rows

    order_date = order_meta.get("date", "")
    total_cost = order_meta.get("total_cost", 0)
    total_msrp = order_meta.get("total_msrp", 0)

    # COGS allocation ratio: cost / MSRP
    cogs_ratio = total_cost / total_msrp if total_msrp > 0 else 0

    pallet_brands = order_meta.get("pallet_brands", {})

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue

        # Find header row
        header_row = None
        header_idx = 0
        for i, row in enumerate(sheet_rows):
            if row and any(
                str(cell).strip().lower() in EXPECTED_HEADERS
                for cell in row if cell
            ):
                header_row = [str(c).strip().lower() if c else "" for c in row]
                header_idx = i
                break

        if not header_row:
            continue

        # Map columns
        col_map = {}
        for j, h in enumerate(header_row):
            if "listing id" in h:
                col_map["listing_id"] = j
            elif "listing title" in h:
                col_map["listing_title"] = j
            elif "category" in h:
                col_map["category"] = j
            elif "product name" in h:
                col_map["product_name"] = j
            elif h == "upc":
                col_map["upc"] = j
            elif h == "asin":
                col_map["asin"] = j
            elif h == "quantity":
                col_map["quantity"] = j
            elif "orig. retail" == h or h == "orig. retail":
                col_map["unit_retail"] = j
            elif "total orig" in h:
                col_map["total_retail"] = j

        # Parse data rows
        for row in sheet_rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def get_val(key, default=None):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return default
                val = row[idx]
                return val if val is not None else default

            listing_id = str(get_val("listing_id", "")).strip()
            product_name = str(get_val("product_name", "")).strip()

            # Skip empty/header rows
            if not product_name or product_name.lower() == "product name":
                continue

            upc_raw = get_val("upc", "")
            upc = str(upc_raw).strip() if upc_raw else ""
            # Clean UPC: remove .0 from float conversion
            if upc.endswith(".0"):
                upc = upc[:-2]
            # Remove non-digit characters for validation but keep original
            upc_clean = re.sub(r'\D', '', upc)
            if len(upc_clean) in (12, 13):
                upc = upc_clean
            elif upc_clean:
                upc = upc_clean  # keep whatever digits we have

            asin = str(get_val("asin", "")).strip() if get_val("asin") else ""

            try:
                quantity = int(float(get_val("quantity", 1)))
            except (ValueError, TypeError):
                quantity = 1

            try:
                unit_retail = float(get_val("unit_retail", 0))
            except (ValueError, TypeError):
                unit_retail = 0.0

            try:
                total_retail = float(get_val("total_retail", 0))
            except (ValueError, TypeError):
                total_retail = unit_retail * quantity

            if total_retail == 0 and unit_retail > 0:
                total_retail = unit_retail * quantity

            # Allocate COGS per unit: (unit_retail * cogs_ratio)
            allocated_cogs_per_unit = unit_retail * cogs_ratio

            # Get brands from pallet mapping
            line_item_brands = pallet_brands.get(listing_id, "")

            rows.append({
                "order_id": order_id,
                "listing_id": listing_id,
                "listing_title": str(get_val("listing_title", "")).strip(),
                "category": str(get_val("category", "")).strip(),
                "product_name": product_name,
                "upc": upc,
                "asin": asin,
                "quantity": quantity,
                "unit_retail": unit_retail,
                "total_retail": total_retail,
                "order_date": order_date,
                "line_item_brands": line_item_brands,
                "allocated_cogs_per_unit": allocated_cogs_per_unit,
            })

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Sync TL manifests to PostgreSQL")
    parser.add_argument("--manifests-dir", default=None, help="Directory containing XLSX manifests")
    parser.add_argument("--orders-json", default=None, help="Path to orders.json")
    args = parser.parse_args()

    # Default paths
    base_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, "data", "techliquidators")

    manifests_dir = args.manifests_dir or os.path.join(data_dir, "order_manifests")
    orders_json = args.orders_json or os.path.join(data_dir, "orders.json")

    print(f"Manifests dir: {manifests_dir}")
    print(f"Orders JSON:   {orders_json}")

    # Load orders metadata
    orders_map = load_orders(orders_json)
    print(f"Loaded {len(orders_map)} orders from orders.json")

    # Find all manifest XLSX files
    xlsx_files = sorted(glob.glob(os.path.join(manifests_dir, "order_manifest_*.xlsx")))
    print(f"Found {len(xlsx_files)} manifest XLSX files")

    if not xlsx_files:
        print("No manifest files found. Exiting.")
        return

    # Connect to DB and create table
    conn = get_db_connection()
    cur = conn.cursor()

    print("Creating table if not exists...")
    cur.execute(CREATE_TABLE_SQL)
    conn.commit()

    # Parse all manifests
    all_rows = []
    for filepath in xlsx_files:
        filename = os.path.basename(filepath)
        # Extract order_id from filename: order_manifest_XXXXXX.xlsx
        order_id = filename.replace("order_manifest_", "").replace(".xlsx", "")

        order_meta = orders_map.get(order_id, {
            "date": "",
            "total_cost": 0,
            "total_msrp": 0,
            "pallet_brands": {},
        })

        rows = parse_manifest_xlsx(filepath, order_id, order_meta)
        print(f"  {filename}: {len(rows)} product rows (order {order_id})")
        all_rows.extend(rows)

    print(f"\nTotal rows to upsert: {len(all_rows)}")

    if not all_rows:
        print("No rows parsed. Exiting.")
        conn.close()
        return

    # Upsert rows
    upsert_sql = """
    INSERT INTO tl_manifest_items (
        order_id, listing_id, listing_title, category, product_name,
        upc, asin, quantity, unit_retail, total_retail,
        order_date, line_item_brands, allocated_cogs_per_unit
    ) VALUES (
        %(order_id)s, %(listing_id)s, %(listing_title)s, %(category)s, %(product_name)s,
        %(upc)s, %(asin)s, %(quantity)s, %(unit_retail)s, %(total_retail)s,
        %(order_date)s, %(line_item_brands)s, %(allocated_cogs_per_unit)s
    )
    ON CONFLICT (order_id, listing_id, product_name, upc)
    DO UPDATE SET
        quantity = EXCLUDED.quantity,
        unit_retail = EXCLUDED.unit_retail,
        total_retail = EXCLUDED.total_retail,
        order_date = EXCLUDED.order_date,
        line_item_brands = EXCLUDED.line_item_brands,
        allocated_cogs_per_unit = EXCLUDED.allocated_cogs_per_unit,
        category = EXCLUDED.category,
        asin = EXCLUDED.asin
    """

    inserted = 0
    errors = 0
    for row in all_rows:
        try:
            cur.execute(upsert_sql, row)
            inserted += 1
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  ERROR inserting row: {e}")
                print(f"    Row: order={row['order_id']}, product={row['product_name'][:50]}, upc={row['upc']}")
            conn.rollback()
            # Try again with this row individually
            try:
                cur.execute(upsert_sql, row)
                inserted += 1
                errors -= 1
            except:
                pass

    conn.commit()

    # Verify
    cur.execute("SELECT COUNT(*) FROM tl_manifest_items")
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT order_id) FROM tl_manifest_items")
    order_count = cur.fetchone()[0]
    cur.execute("SELECT SUM(quantity) FROM tl_manifest_items")
    total_items = cur.fetchone()[0]
    cur.execute("SELECT SUM(total_retail) FROM tl_manifest_items")
    total_msrp = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT upc) FROM tl_manifest_items WHERE upc IS NOT NULL AND upc <> ''")
    unique_upcs = cur.fetchone()[0]

    print(f"\n--- Sync Complete ---")
    print(f"Upserted:      {inserted} rows ({errors} errors)")
    print(f"Total in DB:   {total} manifest rows")
    print(f"Orders:        {order_count}")
    print(f"Total items:   {total_items}")
    print(f"Total MSRP:    ${total_msrp:,.2f}" if total_msrp else "Total MSRP: $0")
    print(f"Unique UPCs:   {unique_upcs}")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
